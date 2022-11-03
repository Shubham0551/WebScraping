import openpyxl
# from ConfigData.ExcelData import *
def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)
def get_columns_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)
def read_data(file, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.cell(row=rownum, column=columnum).value)

def write_data(file, sheetname, rownum, columnum):
    pass

# workbook = openpyxl.load_workbook(excel_url)
# sheet = workbook['list-cities-us-30j']
# row_count=sheet.max_row
# print(row_count)
# column_count=sheet.max_column
# print(column_count)
# for i in range(1, row_count + 1):
#     for j in range(1, column_count+ 1):
#         print(sheet.cell(row=i, column=j).value)